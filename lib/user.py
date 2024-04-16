from openpyxl import load_workbook, Workbook
from random import *

c_name = 1
c_id = 2
c_money = 3
c_lvl = 4
c_exp = 5
c_loss = 6

# 외화
c_dollar = 7
c_yen = 8
c_yuan = 9
c_euro = 10
c_pound = 11
c_real = 12
c_ruble = 13
c_rupee = 14

# 낚시 (미끼)
c_bait = 15

# Coin
c_waves = 16
c_btc = 17
c_eth = 18
c_doge = 19
c_ltc = 20
c_etc = 21
c_lunc = 22
c_sand = 23
c_bnb = 24
c_xrp = 25

# Graps of the Undying
c_graps = 26

# ITEM
c_wisky = 27
c_vitamin = 28
c_bk = 29
c_kc = 30
c_rkc = 31
c_ostrich = 32

default_money = 30000000
default_foreign_currency = 0

wb = load_workbook("./db/userDB.xlsx")
ws = wb.active

def loadFile():
    wb = load_workbook("./db/userDB.xlsx")
    ws = wb.active

def saveFile():
    wb.save("./db/userDB.xlsx")
    wb.close()

#=========================checking==================================
def checkUserNum():
    print("user.py - checkUserNum")
    loadFile()

    count = 0

    for row in range(2, ws.max_row+1):
        if ws.cell(row,c_name).value != None:
            count = count+1
        else:
            count = count
    return count

def checkFirstRow():
    print("user.py - checkFirstRow")
    loadFile()

    print("첫번째 빈 곳을 탐색")

    for row in range(2, ws.max_row + 1):
        if ws.cell(row,1).value is None:
            return row
            break

    _result = ws.max_row+1

    saveFile()

    return _result

def checkUser(_name, _id):
    print("user.py - checkUser")
    print(str(_name) + "<" + str(_id) + ">의 존재 여부 확인")
    print("")

    loadFile()

    userNum = checkUserNum()
    print("등록된 유저수: ", userNum)
    print("")

    print("이름과 고유번호 탐색")
    print("")

    for row in range(2, 3+userNum):
        print(row, "번째 줄 name: ", ws.cell(row,c_name).value)
        print("입력된 name: ", _name)
        print("이름과 일치 여부: ", ws.cell(row, c_name).value == _name)

        print(row,"번째 줄 id: ", ws.cell(row,c_id).value)
        print("입력된 id: ", hex(_id))
        print("고유번호정보와 일치 여부: ", ws.cell(row, c_id).value == hex(_id))
        print("")

        if ws.cell(row, c_name).value == _name and ws.cell(row,c_id).value == hex(_id):
            print("등록된  이름과 고유번호를 발견")
            print("등록된  값의 위치: ",  row, "번째 줄")
            print("")

            saveFile()

            return True, row
            break
        else:
            print("등록된 정보를 탐색 실패, 재탐색 실시")

    saveFile()
    print("발견 실패")

    return False, None

def checkRow():
    loadFile()
    print("첫번쨰 빈 곳을 탐색")
    for row in range(2, ws.max_row + 1):
        if ws.cell(row,1).value is None:
            return row
            break
    _result = ws.max_row+1
    saveFile()

    return _result

#=========================Money==================================
def getMoney(_name,_row):
    print("user.py - getMoney")
    loadFile()

    print(_name, "의 돈을 탐색")

    result = ws.cell(_row, c_money).value
    print(_name,"의 보유 자산: ", result)

    saveFile()

    return result

def getForeignCurrency(_name, _row, _c):
    print("user.py - geForeignCurrency")
    loadFile()

    print(_name, "의 외화를 탐색")

    result = ws.cell(_row, _c).value

    saveFile()

    return result

def remit(sender, sender_row, receiver, receiver_row, _amount):
    print("user.py - remit")
    #loadFile()
    
    print("보내는 사람: ", sender)
    print("받는 사람: ", receiver)
    print("보내는 돈: ", _amount)
    print("")

    modifyMoney(receiver, receiver_row, int(_amount))
    modifyMoney(sender, sender_row, -int(_amount))

    print("")

def modifyMoney(_target, _row, _amount):
    print("user.py - modifyMoney")
    loadFile()
    print(_target, "의 자산데이터 수정")
    print(_target, "의 자산: " + str(ws.cell(_row, c_money).value))
    print("추가할 액수: ", _amount)
    ws.cell(_row, c_money).value += _amount

    print("자산데이터 수정 완료")
    print("수정된", _target, "의 자산: ", ws.cell(_row, c_money).value)
    
    saveFile()

def addLoss(_target, _row, _amount):
    print("user.py - addLoss")
    loadFile()

    print(_target, "의 잃은 돈 추가")
    print(_target, "의 잃은돈: " + str(ws.cell(_row, c_loss).value))
    print("추가할 액수: ", _amount)
    ws.cell(_row, c_loss).value += _amount

    print("잃은 돈 추가 완료")
    print(_target, "의 총 잃은 돈: ", ws.cell(_row, c_loss).value)

    saveFile()

def buyForeignCurrency(_target, _row, _amount, _c):
    print("user.py - buyForeignCurrency")
    loadFile()

    print(_target, "의 외화 자산 데이터 수정")
    print("추가할 액수: ", _amount)

    ws.cell(_row, _c).value += int(_amount)

    print("자산데이터 수정 완료")
    
    saveFile()

def sellForeignCurrency(_target, _row, _amount, _c):
    print("user.py - sellForeignCurrency")
    loadFile()

    print(_target, "의 외화 자산 데이터 수정")
    print("제거할 액수: ", _amount)

    ws.cell(_row, _c).value -= int(_amount)

    print("자산데이터 수정 완료")

    saveFile()
    return

#=========================Level==================================
def levelupCheck(_row):
    print("user.py - levelupCheck")
    loadFile()

    name = ws.cell(_row, c_name).value
    exp = ws.cell(_row, c_exp).value
    lvl = ws.cell(_row, c_lvl).value
    amount_to_up = lvl*lvl + 6*lvl
    count = 0

    print(name,"의 레벨업 조사")
    print(name, "의 현재 레벨: ", lvl, "(", exp, "/", amount_to_up,")")

    if exp >= amount_to_up:
        while(exp >= amount_to_up and exp >= 0):
            print("레벨업에 필요한 경험치 :", amount_to_up)
            print("현재 경험치: ", exp)

            print("충분한 경험치양을 확인")
            ws.cell(_row, c_lvl).value += 1
            count += 1
            print("레벨 데이터 수정")
            ws.cell(_row, c_exp).value -= amount_to_up
            print("경험치 초기화")

            lvl = ws.cell(_row, c_lvl).value
            exp = ws.cell(_row, c_exp).value
            amount_to_up = lvl*lvl + 6*lvl
        return True, lvl
    else:
        return False, lvl

def modifyExp(_row, _amount):
    print("user.py - modifyExp")
    loadFile()

    name = ws.cell(_row, c_name).value
    print(name, "의 경험치 획득량: ", _amount)

    ws.cell(_row, c_exp).value += _amount
    print(name, "현재 경험치: ", ws.cell(_row, c_exp).value)

    saveFile()

#=========================Ranking==================================
def ranking():
    print("user.py - ranking")

    loadFile()

    userRanking =  {}
    userNum = checkUserNum()

    print("등록된 유저수: ", userNum)
    print("")

    print("랭킹 집계중")

    for row in range(2, 2+userNum):
        _name = ws.cell(row, c_name).value
        _lvl = ws.cell(row, c_lvl).value
        userRanking[_name] = _lvl

    print("랭킹 집계 완료")
    a = sorted(userRanking.items(), reverse=True, key=lambda item:item[1])
    result = []
    for items in a:
        result.append(items[0])
        result.append(items[1])
    print(result)
    print("")

    return result

def ranking_money():
    print("user.py - ranking")

    loadFile()

    userRanking =  {}
    userNum = checkUserNum()

    print("등록된 유저수: ", userNum)
    print("")

    print("랭킹 집계중")

    for row in range(2, 2+userNum):
        _name = ws.cell(row, c_name).value
        _money = ws.cell(row, c_money).value
        userRanking[_name] = _money

    print("랭킹 집계 완료")
    a = sorted(userRanking.items(), reverse=True, key=lambda item:item[1])
    result = []
    for items in a:
        result.append(items[0])
        result.append(items[1])
    print(result)
    print("")

    return result

def getRank(_row):
    print("user.py - getRank")
    user = ws.cell(_row, c_name).value
    print(user, "의 랭킹 조사")
    rank = ranking()

    result = int(rank.index(user)/2)+1
    print(user, "의 랭킹: ",result, "위")

    return result

#=========================Account==================================
def Signup(_name, _id):
    print("user.py - signup")

    loadFile()

    _row = checkFirstRow()
    print("첫번째 빈곳: ", _row)
    print("")

    print("데이터 추가 시작")

    ws.cell(row=_row, column=c_name, value=_name)
    print("이름 추가 | ",  ws.cell(_row,c_name).value)
    ws.cell(row=_row, column=c_id, value =hex(_id))
    print("고유번호 추가 | ", ws.cell(_row,c_id).value)

    ws.cell(row=_row, column=c_lvl, value = 1)
    print("초기 레벨 설정 | lvl:", ws.cell(_row,c_lvl).value)
    ws.cell(row=_row, column=c_exp, value = 0)
    print("초기 경험치 설정 | exp:", ws.cell(_row,c_exp).value)

    ws.cell(row=_row, column=c_money, value = default_money)
    print("기본 자금 지급 | ", ws.cell(_row,c_money).value)
    ws.cell(row=_row, column=c_loss, value = 0)
    print("초기 손실 설정 | loss:", ws.cell(_row,c_loss).value)

    ws.cell(row=_row, column=c_dollar, value = default_foreign_currency)
    print("외화 달라 초기화 | ", ws.cell(_row,c_dollar).value)
    ws.cell(row=_row, column=c_yen, value = default_foreign_currency)
    print("외화 엔화 초기화 | :", ws.cell(_row,c_yen).value)
    ws.cell(row=_row, column=c_yuan, value = default_foreign_currency)
    print("외화 위안화 초기화 | ", ws.cell(_row,c_yuan).value)
    ws.cell(row=_row, column=c_euro, value = default_foreign_currency)
    print("외화 유로 초기화 | :", ws.cell(_row,c_euro).value)
    ws.cell(row=_row, column=c_pound, value = default_foreign_currency)
    print("외화 파운드 초기화 | ", ws.cell(_row,c_pound).value)
    ws.cell(row=_row, column=c_real, value = default_foreign_currency)
    print("외화 헤알 초기화 | loss:", ws.cell(_row,c_real).value)
    ws.cell(row=_row, column=c_ruble, value = default_foreign_currency)
    print("외화 루블 초기화 | ", ws.cell(_row,c_ruble).value)
    ws.cell(row=_row, column=c_rupee, value = default_foreign_currency)
    print("외화 루피 초기화 | loss:", ws.cell(_row,c_rupee).value)

    print("코인 초기화 작업을 시작합니다.")
    for index in range(16,30):
        ws.cell(row=_row, column= index, value = 0)
    print("코인 초기화 완료")

    print("")

    saveFile()

    print("데이터 추가 완료")

def DeleteAccount(_row):
    print("user.py - DeleteAccount")
    loadFile()
    print("회원탈퇴 진행")

    print("유저 데이터 삭제")
    ws.delete_rows(_row)

    saveFile()
    
    print("회원탈퇴 완료")

def userInfo(_row):
    loadFile()

    _lvl = ws.cell(_row,c_lvl).value
    _exp = ws.cell(_row,c_exp).value
    _money = ws.cell(_row,c_money).value
    _loss = ws.cell(_row,c_loss).value
    _dolor = ws.cell(_row, c_dollar).value
    _yen = ws.cell(_row, c_yen).value
    _yuan = ws.cell(_row, c_yuan).value
    _euro = ws.cell(_row, c_euro).value
    _pound = ws.cell(_row, c_pound).value
    _real = ws.cell(_row, c_real).value
    _ruble = ws.cell(_row, c_ruble).value
    _rupee = ws.cell(_row, c_rupee).value
    _waves = ws.cell(_row, c_waves).value
    _btc = ws.cell(_row, c_btc).value
    _eth = ws.cell(_row, c_eth).value
    _doge = ws.cell(_row, c_doge).value
    _ltc = ws.cell(_row, c_ltc).value
    _etc = ws.cell(_row, c_etc).value
    _lunc = ws.cell(_row, c_lunc).value
    _sand = ws.cell(_row, c_sand).value
    _bnb = ws.cell(_row, c_bnb).value
    _xrp = ws.cell(_row, c_xrp).value

    print("레벨: ", _lvl)
    print("경험치: ", _exp)
    print("보유자산: ", _money)
    print("잃은 돈: ", _loss)

    saveFile()

    return _lvl, _exp, _money, _loss, _dolor, _yen, _yuan, _euro, _pound, _real, _ruble, _rupee, _waves, _btc, _eth, _doge, _ltc, _etc, _lunc, _sand, _bnb, _xrp

#=========================Coin==================================

def getDollar(_name, _row):
    print("user.py - getDollar")
    loadFile()

    print(_name, "의 달러를 탐색")

    result = ws.cell(_row, c_dollar).value
    print(_name, "의 보유 달러: ", result)

    saveFile()

    return result

def getCoin(_name, _row, _coin):
    print("user.py - getCoin")
    loadFile()

    print(_name, "의 코인을 탐색")

    if _coin == "웨이브":
        result = ws.cell(_row, c_waves).value
        print(_name, "의 보유" + _coin + " : ", result)
    elif _coin == "비트코인":
        result = ws.cell(_row, c_btc).value
        print(_name, "의 보유" + _coin + " : ", result)
    elif _coin == "이더리움":
        result = ws.cell(_row, c_eth).value
        print(_name, "의 보유" + _coin + " : ", result)
    elif _coin == "도지":
        result = ws.cell(_row, c_doge).value
        print(_name, "의 보유" + _coin + " : ", result)
    elif _coin == "라이트":
        result = ws.cell(_row, c_ltc).value
        print(_name, "의 보유" + _coin + " : ", result)
    elif _coin == "이더리움클래식":
        result = ws.cell(_row, c_etc).value
        print(_name, "의 보유" + _coin + " : ", result)
    elif _coin == "루나클래식":
        result = ws.cell(_row, c_lunc).value
        print(_name, "의 보유" + _coin + " : ", result)
    elif _coin == "샌드박스":
        result = ws.cell(_row, c_sand).value
        print(_name, "의 보유" + _coin + " : ", result)
    elif _coin == "비엔비":
        result = ws.cell(_row, c_bnb).value
        print(_name, "의 보유" + _coin + " : ", result)
    elif _coin == "리플":
        result = ws.cell(_row, c_xrp).value
        print(_name, "의 보유" + _coin + " : ", result)
    else:
        return 0

    saveFile()

    return result


def modifyDollar(_target, _row, _amount):
    print("user.py - modifyDollar")
    loadFile()
    print(_target, "의 달라데이터 수정")
    print(_target, "의 달라: " + str(ws.cell(_row, c_dollar).value))
    print("추가할 액수: ", _amount)
    ws.cell(_row, c_dollar).value += _amount

    print("자산데이터 수정 완료")
    print("수정된", _target, "의 자산: ", ws.cell(_row, c_dollar).value)
    
    saveFile()

def modifyCoin(_target, _row, _amount, _coin):
    print("user.py - modifyCoin")
    loadFile()
    print(_target, "의 코인데이터 수정")
    print("추가할 액수: ", _amount)

    if _coin == "웨이브":
        ws.cell(_row, c_waves).value += _amount
        print("자산데이터 수정 완료")
        print("수정된", _target, "의 자산: ", ws.cell(_row, c_waves).value)
    elif _coin == "비트코인":
        ws.cell(_row, c_btc).value += _amount
        print("자산데이터 수정 완료")
        print("수정된", _target, "의 자산: ", ws.cell(_row, c_btc).value)
    elif _coin == "이더리움":
        ws.cell(_row, c_eth).value += _amount
        print("자산데이터 수정 완료")
        print("수정된", _target, "의 자산: ", ws.cell(_row, c_eth).value)
    elif _coin == "도지":
        ws.cell(_row, c_doge).value += _amount
        print("자산데이터 수정 완료")
        print("수정된", _target, "의 자산: ", ws.cell(_row, c_doge).value)
    elif _coin == "라이트":
        ws.cell(_row, c_ltc).value += _amount
        print("자산데이터 수정 완료")
        print("수정된", _target, "의 자산: ", ws.cell(_row, c_ltc).value)
    elif _coin == "이더리움클래식":
        ws.cell(_row, c_eth).value += _amount
        print("자산데이터 수정 완료")
        print("수정된", _target, "의 자산: ", ws.cell(_row, c_eth).value)
    elif _coin == "루나클래식":
        ws.cell(_row, c_lunc).value += _amount
        print("자산데이터 수정 완료")
        print("수정된", _target, "의 자산: ", ws.cell(_row, c_lunc).value)
    elif _coin == "샌드박스":
        ws.cell(_row, c_sand).value += _amount
        print("자산데이터 수정 완료")
        print("수정된", _target, "의 자산: ", ws.cell(_row, c_sand).value)
    elif _coin == "비엔비":
        ws.cell(_row, c_bnb).value += _amount
        print("자산데이터 수정 완료")
        print("수정된", _target, "의 자산: ", ws.cell(_row, c_bnb).value)
    elif _coin == "리플":
        ws.cell(_row, c_xrp).value += _amount
        print("자산데이터 수정 완료")
        print("수정된", _target, "의 자산: ", ws.cell(_row, c_xrp).value)
    else:
        return 0
    
    saveFile()


#=========================ITEM==================================
def checkBait(_row):
    print("user.py - checkBait")
    loadFile()

    print("미끼를 탐색")
    
    result = int(ws.cell(_row, c_bait).value)
    print("보유 미끼 수량: ", result)

    # 미끼가 있을 경우
    if result:
        return False
    
    return True

def checkGrap(_row):
    print("user.py - checkGrap")
    loadFile()
    print("착취 개수 조회")
    result = ws.cell(_row, c_graps).value
    print("조회 완료")
    saveFile()
    return result

def countKC(_row):
    print("user.py - countKC")
    loadFile()
    print("협박 개수 조회")
    result = ws.cell(_row, c_kc).value
    print("조회 완료")
    saveFile()

    return result

def checkKC(_row):
    print("user.py - checkKC")
    loadFile()
    print("협박 개수 조회")
    result = ws.cell(_row, c_kc).value
    print("조회 완료")
    saveFile()
    
    if result:
        return False
    
    return True

def checkRKC(_row):
    print("user.py - checRKC")
    loadFile()
    print("협박 개수 조회")
    result = ws.cell(_row, c_rkc).value
    print("조회 완료")
    saveFile()
    
    if result:
        return True
    
    return False

def checkVita(_row):
    print("user.py - checkVita")
    loadFile()
    print("비타민 조회")
    result = ws.cell(_row, c_vitamin).value
    print("조회 성공")
    saveFile()

    if result:
        return False
    
    return True

def checkBK(_row):
    print("user.py - checkBK")
    loadFile()
    print("바지락칼국수 조회")
    result = ws.cell(_row, c_bk).value
    print("조회 성공")
    saveFile()

    if result:
        return False
    
    return True

def checkWIS(_row):
    print("user.py - checkWIS")
    loadFile()
    print("위스키 조회")
    result = ws.cell(_row, c_wisky).value
    print("조회 성공")
    saveFile()

    if result:
        return False
    
    return True

def checkOE(_row):
    print("user.py - checkOE")
    loadFile()
    print("타조알 조회")
    result = ws.cell(_row, c_ostrich).value
    print("조회 성공")
    saveFile()

    if result:
        return True
    
    return False
    

def threat(_name, _row):
    print("user.py - threat")
    loadFile()
    ws.cell(_row,c_rkc).value += 1
    saveFile()
    print("성공")

def addRKC(_row):
    print("user.py - addRKC")
    loadFile()
    ws.cell(_row,c_rkc).value += 1
    saveFile()

def delRKC(_row):
    print("user.py - delRKC")
    loadFile()
    ws.cell(_row,c_rkc).value -= 1
    saveFile()

def openOE():
    n = randint(1,100)
    if 1 <= n <= 40:
        result = "fail"
        url = "https://i.imgur.com/1In7USq.png"
    elif 41 <= n <= 50:
        result = "낙동강타조알"
        url = "https://i.imgur.com/zzW8cHS.jpg"
    elif 51 <= n <= 57:
        result = "강동완의발포비타민"
        url = "https://i.imgur.com/TmLMVLP.jpg"
    elif 58 <= n <= 64:
        result = "바지락칼국수"
        url = "https://i.imgur.com/ejBBFJm.jpg"
    elif 65 <= n <= 66:
        result = "고승환의 커터칼"
        url = "https://i.imgur.com/UBSG3DT.jpg"
    elif 67 <= n <= 76:
        result = "빡친 타조"
        url = "https://i.imgur.com/YUeLEz6.jpg"
    elif 77 <= n <= 92:
        result = "미끼"
        url = "https://i.imgur.com/P3vA4iV.jpg"
    elif 93 <= n <= 99:
        result = "열구의위스키"
        url = "https://i.imgur.com/luLEIYm.jpg"
    else:
        result = "착취의 손아귀"
        url = "https://i.imgur.com/cFFk2oA.jpg"

    return result, url

def addItem(_row, _item, _amount):
    print("user.py - addItme")
    loadFile()
    if _item == "미끼":
        ws.cell(_row,c_bait).value += _amount
    if _item == "열구의위스키":
        ws.cell(_row,c_wisky).value += _amount*3
    if _item == "강동완의발포비타민":
        ws.cell(_row, c_vitamin).value += _amount*5
    if _item == "바지락칼국수":
        ws.cell(_row, c_bk).value += _amount*2
    if _item == "낙동강타조알":
        ws.cell(_row, c_ostrich).value += _amount
    if _item == "착취의 손아귀":
        ws.cell(_row, c_graps).value += _amount
    if _item == "고승환의 커터칼":
        ws.cell(_row, c_kc).value += _amount
    print("추가 완료")
    saveFile()

def rmItem(_row, _item):
    print("user.py - rmItem")
    loadFile()
    if _item == "미끼":
        ws.cell(_row,c_bait).value -= 1
    if _item == "열구의위스키":
        ws.cell(_row,c_wisky).value -= 1
    if _item == "강동완의발포비타민":
        ws.cell(_row, c_vitamin).value -= 1
    if _item == "바지락칼국수":
        ws.cell(_row, c_bk).value -= 1
    if _item == "낙동강타조알":
        ws.cell(_row, c_ostrich).value -= 1
    if _item == "착취의 손아귀":
        ws.cell(_row, c_graps).value -= 1
    if _item == "고승환의 커터칼":
        ws.cell(_row, c_kc).value -= 1
    print("제거 완료")
    saveFile()